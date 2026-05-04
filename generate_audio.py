#!/usr/bin/env python3
"""KIIP Trang 플래시카드용 Google TTS 음성 생성
- 질문: ko-KR-Chirp3-HD-Algieba (남성)
- 답변: ko-KR-Chirp3-HD-Achernar (여성)
- 출력: audio/q{no}.mp3, audio/a{no}.mp3
- Resumable: 이미 있는 파일은 스킵
"""
import os
import sys
import json
import time
import base64
import urllib.request
import urllib.error
from pathlib import Path
from openpyxl import load_workbook

API_KEY = os.environ.get("GOOGLE_TTS_KEY")
if not API_KEY:
    sys.exit("ERROR: set GOOGLE_TTS_KEY env var first.")

ROOT = Path(__file__).parent
AUDIO_DIR = ROOT / "audio"
AUDIO_DIR.mkdir(exist_ok=True)

# 음성 설정
Q_VOICE = {"languageCode": "ko-KR", "name": "ko-KR-Chirp3-HD-Algieba"}  # 남성, 질문(면접관)
A_VOICE = {"languageCode": "ko-KR", "name": "ko-KR-Chirp3-HD-Achernar"}  # 여성, 답변(짱)
AUDIO_CONFIG = {"audioEncoding": "MP3", "sampleRateHertz": 24000}
ENDPOINT = f"https://texttospeech.googleapis.com/v1/text:synthesize?key={API_KEY}"


def synthesize(text, voice):
    """Google TTS REST API 호출 → MP3 bytes 반환"""
    payload = json.dumps({
        "input": {"text": text},
        "voice": voice,
        "audioConfig": AUDIO_CONFIG,
    }).encode("utf-8")
    req = urllib.request.Request(
        ENDPOINT, data=payload,
        headers={"Content-Type": "application/json; charset=utf-8"},
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        result = json.loads(r.read())
    return base64.b64decode(result["audioContent"])


def load_data():
    """엑셀에서 77개 질문/답변 로드"""
    xlsx = '/Users/dj-mbpm3max/Library/CloudStorage/Dropbox/Claude Code Dropbox/구술시험 기출문제-토픽스쿨/구술시험_모범답안_짱.xlsx'
    wb = load_workbook(xlsx)

    items = []
    # 토픽스쿨 (47개)
    for row in wb["토픽스쿨 (단원별)"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            _, _, _, _, q, a = row
            items.append((q, a))
    # 시대에듀 (30개)
    for row in wb["시대에듀 (모의고사)"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            _, _, _, q, a = row
            items.append((q, a))
    return items


def main():
    items = load_data()
    print(f"Loaded {len(items)} items (질문/답변 쌍)")

    # Job 리스트
    jobs = []
    for i, (q, a) in enumerate(items, 1):
        jobs.append((AUDIO_DIR / f"q{i}.mp3", q, Q_VOICE, "Q"))
        jobs.append((AUDIO_DIR / f"a{i}.mp3", a, A_VOICE, "A"))

    total = len(jobs)
    todo = [j for j in jobs if not j[0].exists()]
    skipped = total - len(todo)
    print(f"Total: {total}  Done: {skipped}  To do: {len(todo)}\n")

    char_count = 0
    done = 0
    failed = []
    t0 = time.time()

    for fname, text, voice, label in todo:
        try:
            mp3 = synthesize(text, voice)
            fname.write_bytes(mp3)
            char_count += len(text)
            done += 1
            if done % 10 == 0 or done == len(todo):
                elapsed = time.time() - t0
                rate = done / elapsed if elapsed > 0 else 0
                eta = (len(todo) - done) / rate if rate else 0
                print(f"  [{done:>3}/{len(todo)}] {fname.name} ({label}, {len(text)}자, {len(mp3)//1024}KB) "
                      f"rate={rate:.1f}/s ETA={eta:.0f}s")
        except urllib.error.HTTPError as e:
            err = e.read().decode("utf-8", errors="replace")[:200]
            print(f"  ✗ HTTP {e.code} {fname.name}: {err}")
            failed.append((fname.name, str(e)))
            if e.code == 429:
                print("  Rate-limited. Sleeping 60s...")
                time.sleep(60)
        except Exception as e:
            print(f"  ✗ {fname.name}: {e}")
            failed.append((fname.name, str(e)))
        time.sleep(0.05)

    print(f"\n=== Done ===")
    print(f"Generated: {done} files, {char_count:,} chars")
    print(f"Cost estimate (Chirp3-HD): ${char_count/1_000_000*16:.4f} (free tier: 1M chars/month)")
    print(f"Total audio files: {len(list(AUDIO_DIR.glob('*.mp3')))}")
    if failed:
        print(f"Failed: {len(failed)} (re-run to retry)")
        for name, err in failed[:5]:
            print(f"  - {name}: {err[:80]}")


if __name__ == "__main__":
    main()
